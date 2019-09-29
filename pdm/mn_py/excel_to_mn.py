#!/usr/bin/python
# coding=utf-8
# Author: jsh

import pymysql, xlwt,xlrd
import openpyxl
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import  make_header

list_in = ['单据日期','公司简称','销售发票号','销售订单号','仓库编码','仓库名称','部门编码','部门名称','业务员编码','销售区域','销售省份','销售城市','客户类型','客户编码','客户名称','最终客户正确编码','最终客户名称','业务类型','产品销售类型-销售、赠送、配套、其他','存货编码','存货名称','项目编码','原币含税单价','销量(盒)','原币价税合计','项目名称','品牌','销售类型编码','规格型号','退补标志','退补数量']
list_out = ['公司简称','出库单号','单据日期','仓库编码','仓库名称','部门编码','部门名称','业务员编码','销售区域','销售省份','销售城市','客户类型','客户编码','客户名称','最终客户正确编码','最终客户名称','业务类型','产品销售类型-销售、赠送、配套、其他','存货编码','存货名称','数量','人份数','项目编码','项目名称','品牌','销售类型编码']

#联接到数据库，并封装循环使用，db是数据库名字
def getCon():
    conn = pymysql.connect(host='172.16.0.181', user='root', password='biosan', db='bidata', charset='utf8')
    return conn.cursor()

def in_sql():
    cur = getCon()
    sql  = 'select * from pdm.invoice_order_mn'
    cur.execute(sql)
    all = cur.fetchall()
    return all

def out_sql():
    cur = getCon()
    sql  = 'select * from pdm.outdepot_order_mn'
    cur.execute(sql)
    all = cur.fetchall()
    return all

def w_excel(res,res2):
    book = openpyxl.Workbook() #新建一个excel
    sheet = book.create_sheet('shouru') #新建一个sheet页
    #写表头
    i = 1
    for header in list_in:
        sheet.cell(1,i,header)
        i+=1
    #写入数据
    for row in range(0,len(res)):
        for col in range(0,len(res[row])):
            sheet.cell(row+2,col+1,res[row][col])
        row+=1
    col+=1

    sheet1 = book.create_sheet('fahuo') #新建一个sheet页
    #写表头
    j = 1
    for header in list_out:
        sheet1.cell(1,j,header)
        j+=1
    #写入数据
    for row in range(0,len(res2)):
        for col in range(0,len(res2[row])):
            sheet1.cell(row+2,col+1,res2[row][col])
        row+=1
    col+=1
    book.save('fhsr'+time.strftime("%Y%m%d", time.localtime())+'.xlsx')
    print("导出成功！")

def sed_email():
    # 发件人
    smtpserver = 'smtp.exmail.qq.com'
    username = 'jiangsunhui@biosan.cn'
    password='JSW19941123aa'
    sender='jiangsunhui@biosan.cn'
    # 收件人
    receiver = ['jiangsunhui@biosan.cn','张梅妮<zhangmeini@biosan.cn>','彭丽<pengli@biosan.cn>',' 李秦秦<liqinqin@biosan.cn>','wangshaojie@biosan.cn']
    # ,'张梅妮<zhangmeini@biosan.cn>','彭丽<pengli@biosan.cn>',' 李秦秦<liqinqin@biosan.cn>'
    # 抄送
    cc = ['jiangsunhui@biosan.cn']

    msg = MIMEMultipart('发货收入bi数据')
    subject = '发货收入bi数据'
    msg['Subject'] = subject
    msg['From'] = 'jiangsunhui@biosan.cn'
    msg['To'] = ";".join(receiver)
    msg['Co'] = ";".join(cc)
    # 设置编码
    # msg['Accept-Language']='zh-CN'
    # msg['Accept-Charset']='ISO-8859-1,utf-8'

    # 构造附件
    file_name = time.strftime("%Y%m%d", time.localtime())+'.xlsx'
    file = r'/home/bidata/pdm/mn_py/fhsr'+time.strftime("%Y%m%d", time.localtime())+'.xlsx'
    sendfile=open(file,'rb').read()
    text_att = MIMEText(sendfile,'base64','utf-8')
    text_att["Content-Type"] = 'application/octet-stream'
    #另一种实现方式
    text_att.add_header('Content-Disposition', 'attachment', filename=('fhsr'+file_name))
    msg.attach(text_att)

    #发送邮件
    smtp = smtplib.SMTP()
    smtp.connect('smtp.exmail.qq.com')
    #我们用set_debuglevel(1)就可以打印出和SMTP服务器交互的所有信息。
    #smtp.set_debuglevel(1)
    smtp.login(username, password)
    smtp.sendmail(sender, receiver, msg.as_string())
    smtp.quit()

if __name__ == '__main__':
    data1 = in_sql()
    data2 = out_sql()
    w_excel(data1,data2)
    sed_email()


