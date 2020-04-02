
# -*- coding: utf-8 -*-
# Author: hkey
from email.header import Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
# from email.utils import parseaddr, formataddr
from email import encoders
from email.mime.base import MIMEBase
import smtplib, os
import pymysql, xlwt
import openpyxl
import time

class CreateExcel(object):
    '''查询数据库并生成Excel文档'''
    def __init__(self, mysql_info):
        self.mysql_info = mysql_info
        self.conn = pymysql.connect(host = self.mysql_info['host'], port = self.mysql_info['port'],
                               user = self.mysql_info['user'], passwd = self.mysql_info['passwd'],
                               db = self.mysql_info['db'], charset='utf8')
        self.cursor = self.conn.cursor()
    def getUserData(self, sql):
        # 查询数据库
        self.cursor.execute(sql)
        table_desc = self.cursor.description
        result = self.cursor.fetchall()
        if not result:
            print('没数据。')
            # 返回查询数据、表字段
        print('数据库查询完毕'.center(30, '#'))
        return result, table_desc

    def writeToExcel(self, data, filename):
        # 生成Excel文档
        # 注意：生成Excel是一列一列写入的。
        result, fileds = data
        # wbk = xlwt.Workbook(encoding='utf-8')
        wbk = openpyxl.Workbook()
        # 创建一个表格
        sheet1 = wbk.create_sheet('装机数据')
        for filed in range(1,len(fileds)+1):
            # Excel插入第一行字段信息
            sheet1.cell(1, filed, fileds[filed-1][0]) # (行，列，数据)

        for row in range(2, len(result)+2):
            # 将数据从第二行开始写入
            for col in range(1, len(fileds)+1):
                sheet1.cell(row, col, result[row-2][col-1]) #(行, 列, 数据第一行的第一列)

        wbk.save(filename)
    def close(self):
        # 关闭游标和数据库连接
        self.cursor.close()
        self.conn.close()
        print('关闭数据库连接'.center(30, '#'))
class SendMail(object):
    '''将Excel作为附件发送邮件'''
    def __init__(self, email_info):
        self.email_info = email_info
        # 使用SMTP_SSL连接端口为465
        self.smtp = smtplib.SMTP_SSL(self.email_info['server'], self.email_info['port'])
        # 创建两个变量
        self._attachements = []
        self._from = ''
    def login(self):
        # 通过邮箱名和smtp授权码登录到邮箱
        self._from = self.email_info['user']
        self.smtp.login(self.email_info['user'], self.email_info['password'])
    # def _format_addr(self, s):
    #     name, addr = parseaddr(s)
    #     return formataddr((Header(name, 'utf-8').encode(), addr))

    def add_attachment(self):
        # 添加附件内容
        # 注意：添加附件内容是通过读取文件的方式加入
        file_path = self.email_info['file_path']
        with open(file_path, 'rb') as file:
            filename = os.path.split(file_path)[1]
            mime = MIMEBase('application', 'octet-stream', filename=filename)
            mime.add_header('Content-Disposition', 'attachment', filename=('gbk', '', filename))
            mime.add_header('Content-ID', '<0>')
            mime.add_header('X-Attachment-Id', '0')
            mime.set_payload(file.read())
            encoders.encode_base64(mime)
            # 添加到列表，可以有多个附件内容
            self._attachements.append(mime)

    def sendMail(self):
        # 发送邮件，可以实现群发
        msg = MIMEMultipart()
        contents = MIMEText(self.email_info['content'], 'plain', 'utf-8')
        msg['From'] = self.email_info['user']
        msg['To'] = self.email_info['to']
        msg['Subject'] = self.email_info['subject']

        for att in self._attachements:
            # 从列表中提交附件，附件可以有多个
            msg.attach(att)
        msg.attach(contents)
        try:
            self.smtp.sendmail(self._from, self.email_info['to'].split(','), msg.as_string())
            print('邮件发送成功，请注意查收'.center(30, '#'))
        except Exception as e:
            print('Error:', e)

    def close(self):
        # 退出smtp服务
        self.smtp.quit()
        print('logout'.center(30, '#'))


if __name__ == '__main__':
    # 数据库连接信息
    mysql_dict = {
        'host': '172.16.0.181',
        'port': 3306,
        'user': 'root',
        'passwd': 'biosan',
        'db': 'pdm'
    }
    # 邮件登录及内容信息
    email_dict = {
    # 手动填写，确保信息无误
        "user": "jiangsunhui@biosan.cn",
        # "to": "jiangsunhui@biosan.cn", # 多个邮箱以','隔开；
       "to": "jiangsunhui@biosan.cn,彭丽<pengli@biosan.cn>,郭冬艳<guodongyan@biosan.cn>", # 多个邮箱以','隔开；
        "server": "smtp.exmail.qq.com",
        'port': 465,    # values值必须int类型
        "username": "jiangsunhui@biosan.cn",
        "password": "JSW19941123aa",
        "subject": "装机档案bi数据",
        "content": '您好，这是本月的数据，请查收！！！',
        'file_path': '装机档案_'+ time.strftime("%Y%m%d", time.localtime()) +'.xlsx'
    }
    sql = 'select * from pdm.crm_account_equipments_gdy;'
    create_excel = CreateExcel(mysql_dict)
    # conn = pymysql.Connect('172.16.0.181','root','biosan','ufdata', charset='utf8')
    # cursor = conn.cursor()
    # conn.commit()
    #    print sql7
    sql_res = create_excel.getUserData(sql)
    create_excel.writeToExcel(sql_res,email_dict['file_path'])
    create_excel.close()
    sendmail = SendMail(email_dict)
    sendmail.login()
    sendmail.add_attachment()
    sendmail.sendMail()
    sendmail.close()

